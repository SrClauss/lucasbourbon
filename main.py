import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import openpyxl
import hashlib
import os
import json
import sys
import queue
import contextlib
import time
import psutil
from datetime import datetime
from login import login
from extractor import search_product
from selenium.common.exceptions import WebDriverException, TimeoutException

# Mantendo os cabeçalhos originais
HEADERS = [
    "code", "name", "pricing", "discount", "pricing_with", "cofins_tax", 
    "cofins_value", "difalst_tax", "difalst_value", "fecop_tax", "fecop_value",
    "icmi_value", "icms_tax", "icms_value", "ipi_tax", "ipi_value", "pis_tax",
    "pis_value", "st_tax", "st_value", "weight", "status", "country_of_origin",
    "customs_tariff", "possibility_to_return", "row_num"
]

header_labels = {
    "code": "Código",
    "name": "Nome",
    "pricing": "Preço",
    "discount": "Desconto",
    "pricing_with": "Preço com Impostos",
    "cofins_tax": "Cofins",
    "cofins_value": "Cofins Valor",
    "difalst_tax": "Difal ST",
    "difalst_value": "Difal ST Valor",
    "fecop_tax": "Fecop",
    "fecop_value": "Fecop Valor",
    "icmi_value": "ICMI Valor",
    "icms_tax": "ICMS",
    "icms_value": "ICMS Valor",
    "ipi_tax": "IPI",
    "ipi_value": "IPI Valor",
    "pis_tax": "PIS",
    "pis_value": "PIS Valor",
    "st_tax": "ST",
    "st_value": "ST Valor",
    "weight": "Peso",
    "status": "Status",
    "country_of_origin": "País de Origem",
    "customs_tariff": "Tarifa Aduaneira",
    "possibility_to_return": "Possibilidade de Devolução"
}

def column_to_index(col_letter: str) -> int:
    """Converte uma letra de coluna do Excel (A, B, C...) para um índice numérico (1, 2, 3...)."""
    index = 0
    for char in col_letter.upper():
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index

class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Atlas Copco Scraper - Completo")
        self.geometry("1200x800")

        if getattr(sys, 'frozen', False):
            self.base_path = os.path.dirname(sys.executable)
        else:
            self.base_path = os.path.dirname(os.path.abspath(__file__))
        self.config_path = os.path.join(self.base_path, 'config.json')
        
        self.login_log_queue = queue.Queue()
        self.scraper_log_queue = queue.Queue()

        self.config = self.load_config()
        if not self.config:
            self.destroy()
            sys.exit()

        self.stop_event = threading.Event()
        self.tasks_queue = queue.Queue()
        self.results_queue = queue.Queue()
        self.unsaved_data = []
        self.worker_threads = []
        self.threads_lock = threading.Lock()
        self.saved_rows = set()
        self.reprocess_rows = set()
        self.total_items = 0
        self.saved_items_count = 0
        default_workers = self.config.get("scraping_settings", {}).get("num_workers", 3)
        self.num_workers_var = tk.IntVar(value=default_workers)
        default_headless = self.config.get("system", {}).get("chrome_options", {}).get("headless", False)
        self.headless_var = tk.BooleanVar(value=default_headless)

        self.create_widgets()
        self.process_login_log_queue()
        self.process_scraper_log_queue()

    def log(self, message):
        """Envia uma mensagem para a área de log de Login/Sistema."""
        self.login_log_queue.put(message)

    def load_config(self):
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
                # Validação essencial
                if 'excel_settings' not in config_data or 'input_columns' not in config_data['excel_settings'] or 'code' not in config_data['excel_settings']['input_columns']:
                     messagebox.showerror("Erro de Configuração", "O arquivo 'config.json' precisa ter a seção 'excel_settings' com 'input_columns' e a chave 'code' definida (ex: \"code\": \"A\").")
                     return None
                return config_data
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar o arquivo de configuração:\n{self.config_path}\n\n{e}")
            return None

    def create_widgets(self):
        # Frame de arquivos
        file_frame = ttk.LabelFrame(self, text="Controle de Arquivos", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Button(file_frame, text="Selecionar Arquivo de Entrada", 
                  command=self.select_input_file).pack(side=tk.LEFT)
        self.input_label = ttk.Label(file_frame, text="Nenhum arquivo selecionado")
        self.input_label.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(file_frame, text="Selecionar Saída", 
                  command=self.select_output_file).pack(side=tk.LEFT)
        self.output_label = ttk.Label(file_frame, text="Nenhum arquivo de saída definido")
        self.output_label.pack(side=tk.LEFT, padx=5)
        
        # Área de logs
        main_log_frame = ttk.Frame(self)
        main_log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        main_log_frame.columnconfigure(0, weight=1)
        main_log_frame.columnconfigure(1, weight=1)
        main_log_frame.rowconfigure(0, weight=1)

        left_log_frame = ttk.LabelFrame(main_log_frame, text="Logs de Login e Sistema", padding=5)
        left_log_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        self.login_log_area = scrolledtext.ScrolledText(left_log_frame, state='disabled', font=('Consolas', 10))
        self.login_log_area.pack(fill=tk.BOTH, expand=True)

        right_log_frame = ttk.LabelFrame(main_log_frame, text="Logs de Raspagem", padding=5)
        right_log_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        self.scraper_log_area = scrolledtext.ScrolledText(right_log_frame, state='disabled', font=('Consolas', 10))
        self.scraper_log_area.pack(fill=tk.BOTH, expand=True)
        
        # Controles de progresso
        progress_frame = ttk.Frame(self)
        progress_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100, orient=tk.HORIZONTAL, mode='determinate')
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        self.progress_label = ttk.Label(progress_frame, text="0/0")
        self.progress_label.pack(side=tk.LEFT, padx=5)
        
        # Controles de execução
        ctrl_frame = ttk.Frame(self)
        ctrl_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.start_btn = ttk.Button(ctrl_frame, text="INICIAR PROCESSAMENTO", command=self.start_process)
        self.start_btn.pack(side=tk.LEFT)
        
        self.stop_btn = ttk.Button(ctrl_frame, text="PARAR E SALVAR", command=self.stop_process, state='disabled')
        self.stop_btn.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(ctrl_frame, text="Workers:").pack(side=tk.LEFT, padx=(20, 2))
        self.workers_spinbox = ttk.Spinbox(ctrl_frame, from_=1, to=20, textvariable=self.num_workers_var, width=5)
        self.workers_spinbox.pack(side=tk.LEFT)
        
        self.headless_check = ttk.Checkbutton(ctrl_frame, text="Rodar em 2º plano (headless)", variable=self.headless_var, onvalue=True, offvalue=False)
        self.headless_check.pack(side=tk.LEFT, padx=(10, 0))
        
        self.status_var = tk.StringVar(value="Pronto para iniciar")
        ttk.Label(ctrl_frame, textvariable=self.status_var, font=('Arial', 10)).pack(side=tk.LEFT, padx=10)
        
        self.eta_var = tk.StringVar(value="ETA: --:--:--")
        ttk.Label(ctrl_frame, textvariable=self.eta_var, font=('Arial', 10, 'italic')).pack(side=tk.RIGHT, padx=10)
        
        self.speed_var = tk.StringVar(value="-- itens/min")
        ttk.Label(ctrl_frame, textvariable=self.speed_var, font=('Arial', 10, 'italic')).pack(side=tk.RIGHT, padx=5)

    def process_log_queue(self, log_queue, area):
        try:
            while True:
                message = log_queue.get_nowait()
                area.config(state='normal')
                area.insert(tk.END, message + "\n")
                area.see(tk.END)
                area.config(state='disabled')
        except queue.Empty:
            pass
        finally:
            self.after(100, lambda: self.process_log_queue(log_queue, area))

    def process_login_log_queue(self):
        self.process_log_queue(self.login_log_queue, self.login_log_area)

    def process_scraper_log_queue(self):
        self.process_log_queue(self.scraper_log_queue, self.scraper_log_area)

    def calculate_file_hash(self, filepath):
        hash_sha256 = hashlib.sha256()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()

    def select_input_file(self):
        file_path = filedialog.askopenfilename(title="Selecione o arquivo Excel de entrada", filetypes=[("Arquivos Excel", "*.xlsx *.xls")])
        if file_path:
            self.input_file = file_path
            self.input_label.config(text=os.path.basename(file_path))
            self.input_hash = self.calculate_file_hash(file_path)
            
            self.output_file = os.path.join(
                os.path.dirname(self.input_file),
                os.path.splitext(os.path.basename(self.input_file))[0] + "_PROCESSADO.xlsx"
            )
            self.output_label.config(text=os.path.basename(self.output_file))
            self.log(f"Arquivo de saída padrão definido para: {self.output_file}")

            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                selected = self.ask_sheet_selection(sheets)
                if selected:
                    self.selected_sheet = selected
                    self.log(f"Planilha selecionada: {selected}")
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível ler o arquivo:\n{str(e)}")

    def ask_sheet_selection(self, sheets):
        popup = tk.Toplevel(self)
        popup.title("Selecionar Planilha")
        popup.geometry("300x300")
        
        tk.Label(popup, text="Selecione a planilha para processar:").pack(pady=10)
        
        selected = tk.StringVar(value=sheets[0])
        for sheet in sheets:
            rb = tk.Radiobutton(popup, text=sheet, variable=selected, value=sheet, padx=20, pady=5)
            rb.pack(anchor='w')
        
        result = []
        def on_ok():
            result.append(selected.get())
            popup.destroy()
        
        tk.Button(popup, text="OK", command=on_ok).pack(pady=10)
        
        popup.grab_set()
        self.wait_window(popup)
        return result[0] if result else None
    
    def select_output_file(self):
        if not hasattr(self, 'input_file'):
            messagebox.showwarning("Aviso", "Selecione um arquivo de entrada primeiro.")
            return
        
        default_name = os.path.splitext(os.path.basename(self.input_file))[0] + "_PROCESSADO.xlsx"
        file_path = filedialog.asksaveasfilename(title="Salvar resultado como", defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")], initialfile=default_name)
        if file_path:
            self.output_file = file_path
            self.output_label.config(text=os.path.basename(file_path))
            self.log(f"Arquivo de saída alterado para: {self.output_file}")

    def check_output_continuity(self):
        self.reprocess_rows.clear()
        self.saved_rows.clear()

        if not hasattr(self, 'output_file') or not os.path.exists(self.output_file):
            self.log("Arquivo de saída não encontrado. Iniciando um novo processamento.")
            return True

        try:
            wb = openpyxl.load_workbook(self.output_file)
            data_sheet_name = self.selected_sheet if hasattr(self, 'selected_sheet') else None
            
            if "Metadata" in wb.sheetnames:
                meta_sheet = wb["Metadata"]
                metadata = {"input_hash": meta_sheet.cell(row=1, column=2).value, "saved_rows_str": meta_sheet.cell(row=4, column=2).value}
                
                if metadata["input_hash"] == self.input_hash:
                    saved_rows_str = metadata.get("saved_rows_str")
                    if saved_rows_str:
                        self.saved_rows = {int(r) for r in str(saved_rows_str).split(',') if r.strip().isdigit()}

                    if data_sheet_name and data_sheet_name in wb.sheetnames:
                        data_sheet = wb[data_sheet_name]
                        status_col_idx = HEADERS.index("status") + 1
                        
                        self.log("Verificando integridade das linhas processadas...")
                        for row_num in range(2, data_sheet.max_row + 1):
                            if not data_sheet.cell(row=row_num, column=status_col_idx).value:
                                self.reprocess_rows.add(row_num)
                                if row_num in self.saved_rows: self.saved_rows.remove(row_num)
                    
                    if self.reprocess_rows: self.log(f"Detectadas {len(self.reprocess_rows)} linha(s) com status vazio (buracos).")

                    response = messagebox.askyesnocancel("Continuar Processamento?", f"Foi encontrado um processamento anterior para este arquivo com {len(self.saved_rows)} itens já salvos.\n{len(self.reprocess_rows)} linha(s) será(ão) reprocessada(s) por estar(em) com status vazio.\n\nDeseja continuar de onde parou?", icon='question')
                    if response is None: return False
                    if response:
                        self.log(f"Continuando processamento. {len(self.saved_rows)} linhas salvas serão ignoradas.")
                        return True
                    else:
                        self.log("Reiniciando o processamento do zero.")
                        self.saved_rows.clear()
                        self.reprocess_rows.clear()
                        return True
                else:
                    if messagebox.askyesno("Arquivo de saída existente", "O arquivo de saída foi gerado a partir de um arquivo de entrada diferente.\nDeseja SOBRESCREVER o arquivo?", icon='warning'):
                        self.log("Sobrescrevendo arquivo de saída.")
                        self.saved_rows.clear()
                        self.reprocess_rows.clear()
                        return True
                    return False
            else:
                if messagebox.askyesno("Arquivo de saída existente", "O arquivo de saída não contém metadados.\nDeseja SOBRESCREVER o arquivo?", icon='warning'):
                    self.log("Sobrescrevendo arquivo de saída sem metadados.")
                    self.saved_rows.clear()
                    self.reprocess_rows.clear()
                    return True
                return False
        except Exception as e:
            self.log(f"Erro ao verificar continuidade: {e}. Assumindo novo processamento.")
            self.saved_rows.clear()
            self.reprocess_rows.clear()
            return True

    def start_process(self):
        if not hasattr(self, 'input_file'):
            messagebox.showwarning("Aviso", "Selecione um arquivo de entrada primeiro!")
            return
        
        if not hasattr(self, 'selected_sheet'):
            messagebox.showwarning("Aviso", "Nenhuma planilha foi selecionada no arquivo de entrada!")
            return

        if not self.check_output_continuity():
            return
        
        if not self.saved_rows and not self.reprocess_rows and hasattr(self, 'output_file') and os.path.exists(self.output_file):
            try:
                os.remove(self.output_file)
                self.log(f"Arquivo de saída '{self.output_file}' removido para reiniciar o processamento.")
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível apagar o arquivo de saída antigo:\n{e}")
                return

        self.saved_items_count = len(self.saved_rows)
        
        self.start_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.workers_spinbox.config(state='disabled')
        self.headless_check.config(state='disabled')
        self.stop_event.clear()
        
        threading.Thread(target=self.run_scraping, daemon=True).start()

    def _worker_manager(self, headless_mode):
        worker_serial_id = 0
        while not self.stop_event.is_set():
            with self.threads_lock:
                self.worker_threads = [t for t in self.worker_threads if t.is_alive()]
                target_workers = self.num_workers_var.get()
                current_workers = len(self.worker_threads)
                
                if current_workers < target_workers:
                    needed = target_workers - current_workers
                    self.login_log_queue.put(f"MANAGER: Solicitando {needed} novo(s) worker(s).")
                    for _ in range(needed):
                        if self.stop_event.is_set(): break
                        worker_serial_id += 1
                        thread = threading.Thread(target=self._scraper_worker, args=(worker_serial_id, headless_mode), daemon=True)
                        thread.start()
                        self.worker_threads.append(thread)
                        time.sleep(15)
            time.sleep(5)

    def _scraper_worker(self, worker_id, headless_mode):
        self.login_log_queue.put(f"[Worker {worker_id}] Iniciando...")
        driver = None
        while not self.stop_event.is_set():
            try:
                if not driver:
                    self.login_log_queue.put(f"[Worker {worker_id}] Tentando fazer login...")
                    driver = login(headless=headless_mode, log_queue=self.login_log_queue)
                    if not driver:
                        self.login_log_queue.put(f"[Worker {worker_id}] ❌ Falha no login. Tentando novamente em 30s.")
                        time.sleep(30)
                        continue
                    self.login_log_queue.put(f"[Worker {worker_id}] ✅ Login bem-sucedido.")

                while not self.stop_event.is_set():
                    try:
                        code, row_num = self.tasks_queue.get(timeout=1)
                        data = search_product(driver, code, worker_id=worker_id, row_num=row_num, log_queue=self.scraper_log_queue)
                        self.results_queue.put(data)
                        self.tasks_queue.task_done()
                    except queue.Empty:
                        continue
                    except (WebDriverException, TimeoutException) as e:
                        self.scraper_log_queue.put(f"🚨 [Worker {worker_id}] Erro no navegador: {type(e).__name__}. Reiniciando.")
                        self.tasks_queue.put((code, row_num))
                        raise
            except Exception as e:
                self.login_log_queue.put(f"🚨 [Worker {worker_id}] Erro crítico, reiniciando: {e}")
                if driver:
                    with contextlib.suppress(Exception): driver.quit()
                driver = None
                time.sleep(15)
        if driver:
            with contextlib.suppress(Exception): driver.quit()
        self.login_log_queue.put(f"[Worker {worker_id}] Finalizado.")

    def run_scraping(self):
        try:
            self.log("\n=== INICIANDO PROCESSAMENTO ===")
            self.tasks_queue = queue.Queue()
            self.results_queue = queue.Queue()

            code_column_letter = self.config['excel_settings']['input_columns']['code']
            self.log(f"Lendo códigos da coluna {code_column_letter}.")
            
            wb_input = openpyxl.load_workbook(self.input_file, read_only=True)
            sheet_input = wb_input[self.selected_sheet]
            
            # ================== INÍCIO DA CORREÇÃO ==================
            all_valid_tasks = []
            self.log(f"Analisando coluna '{code_column_letter}' para encontrar linhas válidas...")
            
            # openpyxl em modo read-only não suporta iteração de colunas diretamente.
            # A forma correta é iterar as linhas e pegar o valor da célula correta.
            code_col_idx = column_to_index(code_column_letter) - 1 # 0-based index para listas

            for row in sheet_input.iter_rows(min_row=2): # Começa da linha 2 para pular o cabeçalho
                # Pega a célula correta da linha atual
                cell = row[code_col_idx]
                
                # Adiciona à lista apenas se a célula tiver um valor e não for apenas espaços em branco
                if cell.value and str(cell.value).strip():
                    all_valid_tasks.append({'code': str(cell.value).zfill(10), 'row_num': cell.row})
            # =================== FIM DA CORREÇÃO ====================

            total_valid_rows = len(all_valid_tasks)
            self.log(f"Encontradas {total_valid_rows} linhas com códigos válidos.")
            self.total_items = total_valid_rows # Define o total correto para a barra de progresso
            
            # Monta a fila de tarefas a serem processadas
            tasks_to_queue = []
            if self.reprocess_rows:
                self.log(f"Priorizando {len(self.reprocess_rows)} linha(s) para reprocessamento.")
                reprocess_tasks = [task for task in all_valid_tasks if task['row_num'] in self.reprocess_rows]
                for task in reprocess_tasks:
                    tasks_to_queue.append((task['code'], task['row_num']))
            
            new_tasks = [task for task in all_valid_tasks if task['row_num'] not in self.saved_rows and task['row_num'] not in self.reprocess_rows]
            for task in new_tasks:
                tasks_to_queue.append((task['code'], task['row_num']))
            
            if tasks_to_queue:
                self.log(f"Total de {len(tasks_to_queue)} tarefas adicionadas à fila.")
                for task in tasks_to_queue:
                    self.tasks_queue.put(task)
            else:
                self.log("Nenhum item novo ou para reprocessar encontrado.")

            wb_input.close()

            # Atualiza a UI com o total correto
            self.progress["maximum"] = self.total_items
            self.progress_label.config(text=f"{self.saved_items_count}/{self.total_items}")
            
            headless_mode = self.headless_var.get()
            manager_thread = threading.Thread(target=self._worker_manager, args=(headless_mode,), daemon=True)
            manager_thread.start()

            start_time = time.time()
            items_processed_session = 0
            while not self.stop_event.is_set():
                try:
                    data = self.results_queue.get(timeout=1)
                    if data:
                        self.unsaved_data.append(data)
                        items_processed_session += 1
                        
                        if len(self.unsaved_data) >= (self.num_workers_var.get() * 2):
                            self.save_data()
                        
                        elapsed = time.time() - start_time
                        if elapsed > 2:
                            speed = items_processed_session / elapsed * 60
                            self.speed_var.set(f"{speed:.1f} itens/min")
                            remaining = self.total_items - (self.saved_items_count + items_processed_session)
                            if speed > 0 and remaining > 0:
                                eta_seconds = remaining / (speed / 60)
                                h, m, s = int(eta_seconds // 3600), int((eta_seconds % 3600) // 60), int(eta_seconds % 60)
                                self.eta_var.set(f"ETA: {h:02d}:{m:02d}:{s:02d}")
                        
                        with self.threads_lock:
                            active_workers = len([t for t in self.worker_threads if t.is_alive()])
                        target_workers = self.num_workers_var.get()
                        self.status_var.set(f"Processando {self.saved_items_count + items_processed_session}/{self.total_items} | Workers: {active_workers}/{target_workers}")

                except queue.Empty:
                    if self.tasks_queue.empty():
                        self.log("Fila de tarefas vazia. Verificando por buracos...")
                        if self._find_and_queue_buracos() == 0:
                            if self.unsaved_data: self.save_data()
                            self.log("Nenhum buraco adicional encontrado. Processamento finalizado.")
                            break
                    time.sleep(0.5)
            
            if self.unsaved_data: self.save_data()
            self.log("\nPROCESSAMENTO CONCLUÍDO." if not self.stop_event.is_set() else "\nProcessamento interrompido.")
            
        except Exception as e:
            self.log(f"\nERRO DURANTE PROCESSAMENTO: {e}")
            import traceback
            self.log(traceback.format_exc())
        finally:
            self.cleanup()
            self.status_var.set("Concluído" if not self.stop_event.is_set() else "Interrompido")
            self.start_btn.config(state='normal')
            self.stop_btn.config(state='disabled')
            self.workers_spinbox.config(state='normal')
            self.headless_check.config(state='normal')
    
    def _find_and_queue_buracos(self):
        buracos = set()
        try:
            if not hasattr(self, 'output_file') or not os.path.exists(self.output_file): return 0
            wb = openpyxl.load_workbook(self.output_file)
            if self.selected_sheet not in wb.sheetnames: return 0
            data_sheet = wb[self.selected_sheet]
            status_col_idx = HEADERS.index("status") + 1
            for row_num in range(2, data_sheet.max_row + 1):
                if not data_sheet.cell(row=row_num, column=status_col_idx).value:
                    buracos.add(row_num)
                    if row_num in self.saved_rows: self.saved_rows.remove(row_num)
        except Exception as e:
            self.log(f"ERRO ao escanear buracos: {e}")
            return 0

        if buracos:
            self.log(f"Detectados {len(buracos)} novo(s) buraco(s) na planilha.")
            wb_input = openpyxl.load_workbook(self.input_file, read_only=True)
            sheet_input = wb_input[self.selected_sheet]
            code_col_idx = column_to_index(self.config['excel_settings']['input_columns']['code'])
            for row_num in sorted(list(buracos)):
                code = sheet_input.cell(row=row_num, column=code_col_idx).value
                if code: self.tasks_queue.put((str(code).zfill(10), row_num))
            wb_input.close()
        return len(buracos)

    def save_data(self):
        if not self.unsaved_data: return
        
        if not hasattr(self, 'output_file'):
             self.log("ERRO CRÍTICO: Tentativa de salvar dados sem um arquivo de saída definido.")
             return

        self.log(f"Salvando lote de {len(self.unsaved_data)} itens em '{os.path.basename(self.output_file)}'...")
        try:
            if os.path.exists(self.output_file):
                wb = openpyxl.load_workbook(self.output_file)
                if self.selected_sheet in wb.sheetnames:
                    data_sheet = wb[self.selected_sheet]
                else:
                    data_sheet = wb.create_sheet(self.selected_sheet)
                    data_sheet.append([header_labels.get(h, h) for h in HEADERS])
            else:
                wb = openpyxl.Workbook()
                data_sheet = wb.active
                data_sheet.title = self.selected_sheet
                data_sheet.append([header_labels.get(h, h) for h in HEADERS])
            
            meta_sheet = wb["Metadata"] if "Metadata" in wb.sheetnames else wb.create_sheet("Metadata")
            
            newly_saved_rows = {item.get('row_num') for item in self.unsaved_data if item.get('row_num')}
            self.saved_rows.update(newly_saved_rows)
            
            meta_sheet['A1'], meta_sheet['B1'] = "Input File Hash", self.input_hash
            meta_sheet['A2'], meta_sheet['B2'] = "Last Processed Row", max(self.saved_rows) if self.saved_rows else 0
            meta_sheet['A3'], meta_sheet['B3'] = "Timestamp", datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            
            meta_sheet['A4'], meta_sheet['B4'] = "Saved Rows", ",".join(map(str, sorted(list(self.saved_rows))))

            for item in self.unsaved_data:
                row_num = item.get('row_num')
                if row_num:
                    row_data = [item.get(h, "") for h in HEADERS]
                    for col_idx, value in enumerate(row_data, start=1):
                        data_sheet.cell(row=row_num, column=col_idx, value=value)
            
            wb.save(self.output_file)
            
            items_saved_count = len(self.unsaved_data)
            self.saved_items_count += items_saved_count
            self.progress_var.set(self.saved_items_count)
            self.progress_label.config(text=f"{self.saved_items_count}/{self.total_items}")
            self.unsaved_data = []
            self.log(f"Lote salvo com sucesso. Total salvo: {self.saved_items_count} linhas.")
        except Exception as e:
            self.log(f"ERRO AO SALVAR: {e}")
            import traceback
            self.log(traceback.format_exc())
    
    def stop_process(self):
        self.stop_event.set()
        self.status_var.set("Finalizando...")
        self.log("\nSolicitação de parada recebida...")

    def save_config(self):
        if not self.config: return
        try:
            self.config.setdefault("scraping_settings", {})["num_workers"] = self.num_workers_var.get()
            self.config.setdefault("system", {}).setdefault("chrome_options", {})["headless"] = self.headless_var.get()
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Aviso: Não foi possível salvar as preferências: {e}")
    
    def cleanup(self):
        self.log("\nSinalizando para workers finalizarem...")
        if not self.stop_event.is_set(): self.stop_event.set()
        with self.threads_lock:
            for thread in self.worker_threads: thread.join(timeout=5)
        try:
            for proc in psutil.process_iter(['pid', 'name']):
                if 'chrome' in proc.info['name'].lower():
                     proc.kill()
        except Exception as e:
            self.log(f"Erro ao limpar processos chrome: {e}")
        self.log("Limpeza concluída.")
    
    def on_closing(self):
        if messagebox.askokcancel("Sair", "Deseja realmente sair?"):
            self.stop_process()
            self.status_var.set("Finalizando... Aguarde.")
            threading.Thread(target=self._perform_cleanup_and_exit, daemon=True).start()

    def _perform_cleanup_and_exit(self):
        self.save_config()
        self.cleanup()
        self.after(0, self.destroy)

if __name__ == "__main__":
    app = Application()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop()
